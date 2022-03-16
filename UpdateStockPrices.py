import openpyxl
import os
import yfinance

# dir = "/Users/michaelsandrik/Documents/Finance/Stock Valuations/"
dir = "/users/michaelsandrik/Library/Mobile Documents/com~apple~CloudDocs/Documents/Finance/Stock Valuations/"

dir_list = os.listdir(dir)

for file in dir_list:
    if file.endswith(".xlsx"):  # only act on Excel files

        tilde = "~"  # check for Excel temp files and skip them
        if tilde in file:
            continue

        print("Opening", file)

        wb = openpyxl.load_workbook(dir + file)
        # ws = wb.active                    # can't depend that workbook will be saved with the correct sheet active
        ws = wb["Valuation"]  # set active worksheet explicitly

        stock_ticker = ws["C4"].value  # get stock ticker from spreadsheet
        print("Updating", stock_ticker)
        print("old price", ws["B16"].value)

        yf_stock = yfinance.Ticker(stock_ticker)  # get stock data from Yahoo Finance
        price = yf_stock.info["regularMarketPrice"]

        ws["B16"].value = price  # update stock price value in spreadsheet

        print("new price", ws["B16"].value, "\n")

        wb.save(dir + stock_ticker + ".xlsx")
